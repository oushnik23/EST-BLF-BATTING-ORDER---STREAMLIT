import streamlit as st
import os
import subprocess
import smtplib
from email.message import EmailMessage
import pandas as pd
import numpy as np
from google.cloud import bigquery
import re

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'D:/Oushnik Sarkar/data-warehousing-prod.json'

st.set_page_config(
    page_title="Parcon BOP Dashboard",
    page_icon="🍃",
    layout="wide")

# ---------------- PAGE DESIGN ---------------- #

st.markdown("""
<style>

/* Top spacing */
.block-container {
    padding-top: 1rem;
}

/* Title */
.main-title {
    text-align: center;
    color: #2E86C1;
    font-size: 40px;
    font-weight: 700;
    margin-top:15px;
    margin-bottom: 7px;
}

/* Subtitle */
.sub-title {
    text-align: center;
    font-size: 28px;
    color: #555;
    margin-top: 5px;
}

/* Divider */
.title-line {
    border-bottom: 1px solid #dcdcdc;
    width: 60%;
    margin: 10px auto 25px auto;
}

/* Email */
.email-label {
    font-size: 24px;
    margin-bottom: 5px;
}

/* 🔥 BUTTON FIX */
div[data-testid="stButton"] > button {
    width: 100%;
    border-radius: 8px;
    border: 1px solid #dcdcdc;
    background-color: #f8f9fa;
    color: #333;
    font-weight: 500;
}

div[data-testid="stButton"] > button:hover {
    border-color: #2E86C1;
    color: #2E86C1;
    background-color: #eef5fb;
}

/* Download */
div[data-testid="stDownloadButton"] > button {
    width: 100%;
    border-radius: 8px;
    border: 1px solid #dcdcdc;
    background-color: #f8f9fa;
}

div[data-testid="stDownloadButton"] > button:hover {
    border-color: #2E86C1;
    color: #2E86C1;
}

/* Section */
.section-header {
    font-size: 20px;
    font-weight: 600;
    margin-top: 20px;
}

</style>

<div class="main-title">Tea CIP (Commodity Intelligence Platform)</div>
<div class="sub-title">EST / BLF Batting Order Position</div>
<div class="title-line"></div>
""", unsafe_allow_html=True)


# ---------------- SETTINGS ---------------- #

working_directory = r"D:\Oushnik Sarkar\Python\STREAMLIT"

modules = [
    {"name":"AS_EST","script":"AS_EST.py","output":"AS_EST.xlsx"},
    {"name":"AS_BLF","script":"AS_BLF.py","output":"AS_BLF.xlsx"},
    {"name":"DO_TR_EST","script":"DO.TR_EST.py","output":"DO_TR_EST.xlsx"},
    {"name":"DO_TR_BLF","script":"DO.TR_BLF.py","output":"DO_TR_BLF.xlsx"},
    {"name":"CATP","script":"CA.TP.py","output":"CATP.xlsx"},
    {"name":"AS_ORTH","script":"AS_ORTH.py","output":"AS_ORTH.xlsx"},
    {"name":"AS_ORTH_EST","script":"AS_ORTH_EST.py","output":"AS_ORTH_EST.xlsx"},
    {"name":"AS_ORTH_BLF","script":"AS_ORTH_BLF.py","output":"AS_ORTH_BLF.xlsx"}
]

combined_output = "EST BLF BATTING ORDER UPTO SALE 13_updated.xlsx"

# ---------------- EMAIL UI ---------------- #

st.markdown('<div class="email-label">✉️ Enter Your Email Address</div>', unsafe_allow_html=True)

receiver_email = st.text_input(
    "",
    placeholder="xxxxx@email.com"
)

# ---------------- EMAIL FUNCTION ---------------- #
def send_email(file_path, receiver_email):

    if not os.path.exists(file_path):
        st.error("❌ File not found. Run process first.")
        return

    try:
        sender_email = "website@parcon.in"

        msg = EmailMessage()
        msg["Subject"] = "EST BLF BATTING ORDER"
        msg["From"] = sender_email
        msg["To"] = receiver_email

        msg.set_content("Please find the attached file.")

        with open(file_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(file_path)
            )

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(sender_email, "xusq bocs tgrk kwig")
            server.send_message(msg)

        st.success(f"Email sent to {receiver_email}",icon="✅")
        #st.balloons()

    except Exception as e:
        st.error(str(e))

# ---------------- RUN SCRIPT ---------------- #
def run_script(module):

    script_path = os.path.join(working_directory, module["script"])
    output_path = os.path.join(working_directory, module["output"])

    try:
        subprocess.run(["python", script_path], check=True)
        st.session_state[module["name"]] = output_path
        st.success(f"{module['name']} completed")

    except:
        st.error(f"{module['name']} failed")

# ---------------- COMBINE FUNCTION ---------------- #

from openpyxl import load_workbook, Workbook
from copy import copy

def combine_excels():

    wb_new = Workbook()
    wb_new.remove(wb_new.active)  # remove default sheet

    for module in modules:

        file_path = os.path.join(working_directory, module["output"])

        if os.path.exists(file_path):

            wb = load_workbook(file_path)
            ws = wb.active

            ws_new = wb_new.create_sheet(title=module["name"])

            # ---------------- COPY DATA + STYLE ---------------- #
            for row in ws.iter_rows():
                for cell in row:

                    new_cell = ws_new.cell(
                        row=cell.row,
                        column=cell.column,
                        value=cell.value
                    )

                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.alignment = cell.alignment.copy()

            # ---------------- COPY COLUMN WIDTH ---------------- #
            for col in ws.column_dimensions:
                ws_new.column_dimensions[col].width = ws.column_dimensions[col].width

            # ---------------- COPY MERGED CELLS  ---------------- #
            for merged_range in ws.merged_cells.ranges:
                ws_new.merge_cells(str(merged_range))

    wb_new.save(combined_output)

    return True

# ---------------- MODULE UI ---------------- #
#st.markdown("### ⚙️ Individual Modules")

st.markdown('<div class="section-header">⚙️ Individual Modules</div>', unsafe_allow_html=True)
st.markdown("<br>", unsafe_allow_html=True)

cols = st.columns(4)

for i, module in enumerate(modules):

    with cols[i % 4]:

        if st.button(f"🍃 {module['name']}"):
            run_script(module)

        if module["name"] in st.session_state:

            file_path = st.session_state[module["name"]]

            if os.path.exists(file_path):

                with open(file_path, "rb") as f:
                    st.download_button(label="📥 Download",data=open(file_path, "rb").read(),file_name=module["output"],
	mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                if st.button("📧 Email", key=module["name"] + "_email"):

                    if receiver_email and "@" in receiver_email:
                        send_email(file_path, receiver_email)
                    else:
                        st.warning("Enter valid email")

# ---------------- COMBINED PROCESS ---------------- #
st.markdown("---")

if st.button("▶ Run Batting Order Process"):

    progress = st.progress(0)
    status = st.empty()

    for i, module in enumerate(modules):

        script_path = os.path.join(working_directory, module["script"])
        status.info(f"Running {module['script']}...")

        subprocess.run(["python", script_path])

        progress.progress((i + 1) / len(modules))

    # 🔥 Combine files
    status.info("Combining Excel files...")
    success = combine_excels()

    if success:
        status.success("🎉 Combined file created!")
    else:
        st.error("❌ No files to combine")

    #st.balloons()

# ---------------- DOWNLOAD COMBINED ---------------- #
if os.path.exists(combined_output):

    with open(combined_output, "rb") as f:
        data = f.read()

    st.download_button(
        label="📥 Download Combined File",
        data=data,
        file_name=os.path.basename(combined_output),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------------- EMAIL COMBINED ---------------- #
if st.button("✉️ Send Combined Email"):

    if receiver_email and "@" in receiver_email:
        send_email(combined_output, receiver_email)
    else:
        st.warning("Enter valid email")

# -------------------------------- Groq AI -------------------------------- #

st.markdown("---")

from groq import Groq

client = Groq(api_key="api_key")

# 1.Detect Query Type

def is_simple_query(user_input):

    text = user_input.lower()

    # 🔥 NEW: trend detection
    if "last" in text and ("year" in text or "season" in text):
        return "trend"

    keywords = ["top", "highest", "lowest", "compare", "rank"]

    if any(word in text for word in keywords):
        return "aggregation"

    return "ai"

def build_grade_query(user_input):

    text = user_input.lower()

    # -------- Garden -------- #
    garden = get_garden_name(user_input)

    # -------- Centre -------- #
    centre_condition = 'Centre IN ("KOL","GUW","SIL")'

    if "kolkata" in text or "kol" in text:
        centre_condition = 'Centre = "KOL"'
    elif "guwahati" in text or "guw" in text:
        centre_condition = 'Centre = "GUW"'
    elif re.search(r"\bsiliguri\b", text) or (" sil " in f" {text} "):
        centre_condition = 'Centre = "SIL"'

    # -------- YEAR LOGIC (same as yours) -------- #
    end_year = 2025
    years = None

    year_match = re.search(r"\b(20\d{2})\b", text)
    if year_match:
        end_year = int(year_match.group(1))
        years = 1

    last_match = re.search(r"last\s+(\d+)", text)
    if last_match:
        years = int(last_match.group(1))

    if years is None:
        years = 2

    start_year = end_year - (years - 1)

    # -------- SQL -------- #
    query = f"""
    SELECT
        Season,
        GardenMDM,SubTeaType,
        GradeMDM,
        coalesce(SUM(Value)) AS Total_Value,
        SUM(TotalWeight) AS Sold_Qty,
        ROUND(SAFE_DIVIDE(SUM(Value), SUM(TotalWeight)),2) AS AvgPrice

    FROM `data-warehousing-prod.EasyReports.SaleTransactionView`

    WHERE Season BETWEEN {start_year} AND {end_year}
        AND EstBlf = "EST"
        AND Category = "CTC"
        AND {centre_condition}
        AND LOWER(GardenMDM) LIKE '%{garden}%'

    GROUP BY Season, GardenMDM, GradeMDM,SubTeaType

    Having Season is not null
    """

    return query

# 2.Fast Rule-Based SQL
import re

def build_fast_query(user_input):

    text = user_input.lower()

    # ---------------- DEFAULTS ---------------- #
    top_n = 10
    offset = 0
    order_by = "Sold_Qty DESC"
    area_condition = ""

    # ---------------- AREA DETECTION ---------------- #
    text = f" {text} "
    text_clean = f" {text} "
    
    if re.search(r"\bassam\b", text_clean):
        area_condition = 'Area = "AS"'
    elif re.search(r"\b(dooars|do|tr)\b", text_clean):
        area_condition = 'Area IN ("DO","TR")'
    elif re.search(r"\b(ca|tp)\b", text_clean):
        area_condition = 'Area IN ("CA","TP")'
        
    
    # -------- Detect Centre -------- #
    centre_condition = 'Centre IN ("KOL","GUW","SIL")'

    if "kolkata" in text or "kol" in text:
        centre_condition = 'Centre = "KOL"'
    elif "guwahati" in text or "guw" in text:
        centre_condition = 'Centre = "GUW"'
    elif "Siliguri" in text:
        centre_condition = 'Centre = "SIL"'

  
# ---------------- METRIC ---------------- #
    if "price" in text or "avg" in text:
        order_by = "AvgPrice DESC"
    elif "lowest" in text:
        order_by = "Sold_Qty ASC"
    elif "qty" in text or "quantity" in text:
        order_by = "Sold_Qty DESC"

    # ---------------- LOWEST CASE ---------------- #
    if "lowest" in text and ("price" in text or "avg" in text):
        order_by = "AvgPrice ASC"

    # ---------------- TOP N ---------------- #
    match_top = re.search(r"top\s+(\d+)", text)
    if match_top:
        top_n = int(match_top.group(1))

    # ---------------- HIGHEST / LOWEST DEFAULT ---------------- #
    elif "highest" in text or "lowest" in text:
        top_n = 1

    # ---------------- 2nd / 3rd / nth ---------------- #
    match_rank = re.search(r"(\d+)(st|nd|rd|th)", text)
    if match_rank:
        rank = int(match_rank.group(1))
        offset = rank - 1
        top_n = 1

    # ---------------- RANGE (top 10 to 20) ---------------- #
    match_range = re.search(r"top\s+(\d+)\s*(to|-)\s*(\d+)", text)
    if match_range:
        start = int(match_range.group(1))
        end = int(match_range.group(3))
        offset = start - 1
        top_n = end - start + 1
        
    cutoff = None

    match_cutoff = re.search(r"(?:cut[- ]?off|above|greater than|more than)\s*(\d+)", text)
    if match_cutoff:
        cutoff = int(match_cutoff.group(1))

    # ---------------- SQL ---------------- #
    query = f"""
    SELECT
        GardenMDM,
        SUM(TotalWeight) AS Sold_Qty,
        ROUND( SAFE_DIVIDE(SUM(Value) , SUM(TotalWeight)) ,2) AS AvgPrice
        
    FROM `data-warehousing-prod.EasyReports.SaleTransactionView`

    WHERE Season = 2025
        AND EstBlf = "EST"
        AND Category = "CTC"
        AND Centre IN ("KOL","GUW","SIL")
        {f"AND {area_condition}" if area_condition else ""}

    GROUP BY GardenMDM
    
    {f"HAVING SUM(TotalWeight) > {cutoff}" if cutoff else ""}

    ORDER BY {order_by}

    LIMIT {top_n}
    OFFSET {offset}
    """

    return query

#----------------BuyerGroup----------------#

def build_buyer_query(user_input):

    text = user_input.lower()
    garden = get_garden_name(user_input)

    # 🔥 TOP N detection (reuse your logic)
    top_n = 10
    match_top = re.search(r"top\s+(\d+)", text)
    if match_top:
        top_n = int(match_top.group(1))
        
        years = 2
        
        match = re.search(r"last\s+(\d+)", text)
        if match:
            years = int(match.group(1))

        end_year = 2025
        years = None
              
        year_match = re.search(r"\b(20\d{2})\b", text)
        if year_match:
            end_year = int(year_match.group(1))
            years = 1
            
        match = re.search(r"last\s+(\d+)", text)
        if match:
            years = int(match.group(1))
            
        if years is None:
            years = 2
        
        start_year = end_year - (years - 1)
    
    if "avg" in text or "price" in text:
        rank_metric = "SAFE_DIVIDE(SUM(Total_Value), SUM(Buyer_Qty)) DESC"
    elif "qty" in text or "quantity" in text:
        rank_metric = "SUM(Buyer_Qty) DESC"
    else:
    # default
        rank_metric = "SUM(Buyer_Qty) DESC"

    query = f"""
    WITH t1 AS (
    SELECT CASE WHEN CAST(SUBSTRING(FinYear,1,4) AS INT64) = Season THEN Season ELSE 0 END AS FYear,

        BuyerGroup,
        GardenMDM,

        SUM(Value) AS Total_Value,
        SUM(TotalWeight) AS Buyer_Qty

    FROM `data-warehousing-prod.EasyReports.SaleTransactionView`

    WHERE Season BETWEEN {start_year} AND {end_year} And EstBlf = "EST" AND Category = "CTC" AND Centre IN ("KOL","GUW","SIL") 
    AND LOWER(GardenMDM) LIKE '%{garden}%'

    GROUP BY FYear, BuyerGroup, GardenMDM, FinYear, Season

    HAVING BuyerGroup <> "" AND FYear <> 0)

    SELECT FYear,GardenMDM,BuyerGroup,
        
        Round(SUM(Buyer_Qty),0) AS Buyer_Qty,
        Round(SUM(Total_Value)/SUM(Buyer_Qty),2) as Avg,

    DENSE_RANK() OVER (PARTITION BY FYear, GardenMDM ORDER BY {rank_metric} ) AS Buyer_Rank

    FROM t1

    GROUP BY FYear, BuyerGroup, GardenMDM

    QUALIFY Buyer_Rank <= {top_n}"""

    return query

# 3.Groq AI (for complex queries)

def extract_garden_name(user_input):

    text = user_input.lower()

    stopwords = ["give", "me", "show", "tell", "find", "get",
        "qty", "quantity", "avg", "average", "price",
        "top", "highest", "lowest", "compare", "rank",
        "last", "year", "years", "season",
        "kolkata", "kol", "guwahati", "guw", "siliguri", "sil",
        "assam", "dooars", "tr", "ca", "tp",
        "for", "and", "in", "of", "by", "upto", "saleno","from", "to", "till", "upto","saleno", "sale",
        "grade", "gradewise", "gradewise", "gradewise", "wise","performance","with","buyers"]

    text = re.sub(r"\d+", "", text)
    text = re.sub(r"[^a-zA-Z\s]", " ", text)
    text = re.sub(r"\b(give|me|show|tell|get|find|with|performance|report|analysis|buyers)\b", "", text)
    text = re.sub(r"\bas per\b", "", text)
    text = re.sub(r"\bbased on\b", "", text)
    text = re.sub(r"\baccording to\b", "", text)

    words = text.split()
    filtered_words = [w for w in words if w not in stopwords]

    if len(filtered_words) == 1:
        return filtered_words[0]

    if len(filtered_words) >= 1:
        return " ".join(filtered_words)

    return None  

def extract_garden_name_ai(user_input):

    prompt = f"""
Extract ONLY the garden name from this query.

Rules:
- Return ONLY one word or name
- No explanation
- Ignore words like qty, avg, price, last, year, kolkata, siliguri etc.

Query:
{user_input}
"""

    response = client.chat.completions.create(
        messages=[{"role": "user", "content": prompt}],model="llama-3.1-8b-instant"
        #model="llama3-70b-8192"
    )

    return response.choices[0].message.content.strip().lower()

def get_garden_name(user_input):

    garden = extract_garden_name(user_input)

    # 🔥 fallback to AI if weak
    if garden is None or len(garden) < 3:
        garden = extract_garden_name_ai(user_input)

    return garden

# ---------------- GARDEN TREND QUERY ---------------- #
def build_garden_trend_query(user_input):

    text = user_input.lower()
    
    # -------- Detect garden name -------- #
    def extract_garden_name(text):

    # remove known keywords
        text = re.sub(
            r"\b(qty|quantity|avg|average|price|for|last|years|year|and|upto|saleno|kolkata|kol|guwahati|guw|siliguri)\b",
            "",
            text)

    # remove numbers
        text = re.sub(r"\d+", "", text)
        
        text = re.sub(r"[^a-zA-Z\s]", "", text)

    # clean extra spaces
        text = text.strip()

        return text

    # -------- Detect years -------- #
    text = user_input.lower()
 
    # -------- Detect Centre -------- #
    centre_condition = 'Centre IN ("KOL","GUW","SIL")'

    if ("kolkata" in text and "guwahati" in text) or ("kol" in text and "guw" in text):
        centre_condition = 'Centre IN ("KOL","GUW")'
    elif ("kolkata" in text and "siliguri" in text) or ("kol" in text and "sil" in text):
        centre_condition = 'Centre IN ("KOL","SIL")'
    elif "kolkata" in text or "kol" in text:
        centre_condition = 'Centre = "KOL"'
    elif "guwahati" in text or "guw" in text:
        centre_condition = 'Centre = "GUW"'
    elif re.search(r"\bsiliguri\b", text) or (" sil " in f" {text} "):
        centre_condition = 'Centre = "SIL"'
    
    # -------- AREA DETECTION -------- #
    area_condition = None
    is_area_mode = False

    area_condition = ""

    text_clean = f" {text} "

    if re.search(r"\bassam\b", text_clean):
        area_condition = 'Area = "AS"'
        is_area_mode = True
    elif re.search(r"\b(dooars|do|tr)\b", text_clean):
        area_condition = 'Area IN ("DO","TR")'
        is_area_mode = True
    elif re.search(r"\b(ca|tp)\b", text_clean):
        area_condition = 'Area IN ("CA","TP")'
        is_area_mode = True
    
    #SaleNo Logic-----------------------------------------------
          
    start_sale = 14
    end_sale = 66
    
    # 🔥 Case 1: "upto sale 48" / "till sale 48"
    upto_match = re.search(r"(upto|till)\s*(sale|saleno)\s*(\d+)", text)
    
    # 🔥 Case 2: "sale 48" / "for sale 48"
    exact_match = re.search(r"(?:for\s*)?(sale|saleno)\s*(\d+)", text)
    
    # 🔥 Case 3: range "sale 10 to 20"
    range_match = re.search(r"(?:from\s*)?(?:sale|saleno)\s*(\d+)\s*(?:to|-)\s*(\d+)", text)
    
    if range_match:
        start_sale = int(range_match.group(1))
        end_sale = int(range_match.group(2))
        
    elif upto_match:
        end_sale = int(upto_match.group(3))   # 14 → X

    elif exact_match:
        sale_no = int(exact_match.group(2))
        start_sale = sale_no
        end_sale = sale_no
 
    garden = get_garden_name(user_input)
    garden = garden.strip()
    years = 2
    
    match = re.search(r"last\s+(\d+)", text)
    if match:
        years = int(match.group(1))

    end_year = 2025
    years = None
          
    year_match = re.search(r"\b(20\d{2})\b", text)
    if year_match:
        end_year = int(year_match.group(1))
        years = 1
        
    match = re.search(r"last\s+(\d+)", text)
    if match:
        years = int(match.group(1))
        
    if years is None:
        years = 2
    
    start_year = end_year - (years - 1)

    # -------- SQL -------- #
    if is_area_mode:

        query = f"""
    SELECT
    CASE WHEN CAST(SUBSTRING(FinYear,1,4) AS INT64) = Season THEN Season ELSE 0 END AS FYear,
    Case 
    when Area = "AS" and Centre IN("KOL","GUW") then "AS"
    when Area IN ("DO","TR") and Centre IN("KOL","SIL") then "DO/TR" 
    when Area IN ("CA","TP") and Centre IN("KOL","GUW") then "CA/TP" else "" end as AreaAlies,
    ROUND(SUM(TotalWeight)/100000,2) AS Sold_Qty_LKGS,
    ROUND(SAFE_DIVIDE(SUM(Value), SUM(TotalWeight)),2) AS AvgPrice
        
    FROM `data-warehousing-prod.EasyReports.SaleTransactionView`

    WHERE Season BETWEEN {start_year} AND {end_year} AND EstBlf = "EST" AND Category = "CTC"
    AND {centre_condition} AND {area_condition}
    AND IF(SaleNo>=1 AND SaleNo<=13, 53+SaleNo, SaleNo) BETWEEN {start_sale} AND {end_sale}

    GROUP BY FYear,AreaAlies

    HAVING FYear <> 0

    ORDER BY FYear DESC
    """
    else:
        query = f"""
    SELECT
    Case when CAST(substring(FinYear,1,4) as INT64) = Season then Season else 0 end as FYear,
    SellerGroup,
    GardenMDM,
    SUM(TotalWeight) AS Sold_Qty,
    ROUND(SAFE_DIVIDE(SUM(Value), SUM(TotalWeight)),2) AS AvgPrice
        
    FROM `data-warehousing-prod.EasyReports.SaleTransactionView`

    WHERE Season BETWEEN {start_year} AND {end_year}
        AND EstBlf = "EST"
        AND Category = "CTC"
        AND {centre_condition}
        AND LOWER(GardenMDM) = '{garden}'
        AND IF(SaleNo>=1 AND SaleNo<=13, 53+SaleNo, SaleNo) BETWEEN {start_sale} AND {end_sale}

    GROUP BY FYear,Season,SellerGroup, GardenMDM
    
    Having FYear<>0

    ORDER BY Season DESC
    """

    return query


def generate_ai_sql(user_input):

    prompt = f"""
You are a BigQuery SQL expert.

Table:
data-warehousing-prod.EasyReports.SaleTransactionView

Rules:
- Always use Season = 2025
- EstBlf = 'EST'
- Category = 'CTC'
- Centre IN ("KOL","GUW","SIL")
- Limit 1000 rows

User Query:
{user_input}

Return ONLY SQL query.
"""

    response = client.chat.completions.create(
        messages=[{"role": "user", "content": prompt}],
        model="llama3-70b-8192"
    )

    return response.choices[0].message.content.strip()

# 4.Clean SQL     
def clean_sql(sql):
    return sql.replace("```sql", "").replace("```", "").strip()

# 5.BigQuery Runner
def run_query(sql):
    client = bigquery.Client()
    return client.query(sql).to_dataframe()

# Streamlit Integration
st.markdown('<div class="section-header">🤖 Type Your Query</div>', unsafe_allow_html=True)

user_query = st.text_input(
    "",
    placeholder="e.g. Borjan, AS garden, top 10 gardens by price")

def is_grade_query(user_input):
    text = user_input.lower()
    
    return any(word in text for word in ["grade", "gradewise", "grade wise"]) and get_garden_name(user_input) is not None

def is_buyer_query(user_input):
    text = user_input.lower()
    return any(word in text for word in ["buyer", "buyergroup"])

if st.button("🚀 Run Smart Query"):

    if user_query:

        with st.spinner("Processing..."):

            try:
                # 🔥 Decide path
                text = user_query.lower()
                
                if is_grade_query(user_query):
                    sql = build_grade_query(user_query)
                    st.info("📊 Grade-wise Mode")
                    
                elif is_buyer_query(user_query):   # 🔥 ADD THIS HERE
                    sql = build_buyer_query(user_query)
                    st.info("👤 Buyer Ranking Mode")
                
                elif re.search(r"\b(20\d{2})\b", text) or ("last" in text and "year" in text):
                    sql = build_garden_trend_query(user_query)
                    st.info("📊 Garden Trend Mode")
                    
                    st.session_state["is_area_mode"] = "Sold_Qty_LKGS" in sql
                    
                elif is_simple_query(user_query):
                    sql = build_fast_query(user_query)
                    st.info("⚡ Fast Mode")
                    
                else:
                    sql = generate_ai_sql(user_query)
                    sql = clean_sql(sql)
                    st.info("🧠 AI Mode")
                
                # 🔍 Show SQL
                st.code(sql, language="sql")

                df = run_query(sql)
                st.session_state["df"] = df
                st.session_state["user_query"] = user_query
                
                if df.empty:
                    st.warning("⚠️ No data return")
                    st.stop()
            except Exception as e:
                st.error(str(e))
    else:
        st.warning("Please enter query")

# ================= DISPLAY SECTION ================= #
                
if "df" in st.session_state:

    df = st.session_state["df"]
    user_query = st.session_state["user_query"]

    view_type = st.radio("📊 View Type", ["Normal", "Tabular Report"])

    if view_type == "Normal":
        st.success("✅ Data fetched")
        st.dataframe(df, use_container_width=True)
        
        if st.session_state.get("is_area_mode"):
            st.caption("📌 Note: Quantity is in Lakh Kgs")

    else:
        # 🔥 ONLY Grade Query → Pivot
        if is_grade_query(user_query):

            st.info("📊 Grade-wise Pivot View")
            
            summary=df.groupby(['Season','GardenMDM','SubTeaType','GradeMDM'],
                               as_index=False).agg({'Sold_Qty':'sum','Total_Value':'sum'})
            summary['Avg']=summary['Total_Value'] / summary['Sold_Qty']
            #summary.drop(['Total_Value'],inplace=True,axis=1)
            
            detail = summary.copy()
                        
            subtotal = summary.groupby(
        ['Season','GardenMDM', 'SubTeaType'],as_index=False).agg({'Sold_Qty': 'sum','Total_Value': 'sum'})
            subtotal['Avg'] = subtotal['Total_Value'] / subtotal['Sold_Qty']
            subtotal['GradeMDM'] = "SUBTOTAL"
                        
            grand = summary.groupby(['Season','GardenMDM'],as_index=False).agg({
        'Sold_Qty': 'sum','Total_Value': 'sum'})
            grand['Avg'] = grand['Total_Value'] / grand['Sold_Qty']
            grand['SubTeaType'] = "ALL"
            grand['GradeMDM'] = "GRAND TOTAL"
                        
            final_df = pd.concat([detail, subtotal, grand], ignore_index=True)
            
            final_df = final_df.drop(columns=['Total_Value'])
                 
                                         
            grade_sequence = ["BOPL", "BPS", "BOP", "BOPSM", "BP", "PF", "OF", "PD", "D", "CD",
    "BOPL1", "BPS1", "BOP1", "BOPSM1", "BP1", "PF1", "OF1", "PD1", "D1", "CD1","SUBTOTAL","GRAND TOTAL"]
                        
            grade_order = {g: i for i, g in enumerate(grade_sequence)}
            ps_order = {'PRIMARY': 0, 'SECONDARY': 1, 'ALL': 2}
            
            final_df['ps_rank'] = final_df['SubTeaType'].map(ps_order).fillna(999)
            final_df["grade_sort"] = final_df["GradeMDM"].map(grade_order)
            
            final_df.loc[final_df["GradeMDM"] == "Subtotal", "grade_sort"] = 998
            final_df.loc[final_df["GradeMDM"] == "Grand Total", "grade_sort"] = 999
            
                       
            final_df = final_df.sort_values(by=["GardenMDM", "Season", "ps_rank", "grade_sort"],
    ascending=[True, False, True, True])
            
            final_df = final_df.drop(columns=["grade_sort","ps_rank"])
            final_df["Avg"] = final_df["Avg"].round(0).astype(int)
                                   
            st.success("✅ Data fetched")
            st.dataframe(final_df, use_container_width=True)
